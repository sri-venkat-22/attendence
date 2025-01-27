import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os


# loading data sheet
data = openpyxl.load_workbook('/Users/srivenkatreddy/Documents/attendence.xlsx')
#choseing sheet
sheet = data['Sheet1']

r = sheet.max_row
c = sheet.max_column

# list of students who need to be reminded of attendance and lack of attendance
std_mail_id = []
lack_attd = ""
l3 =[]


FROM_EMAIL = os.getenv('FROM_EMAIL','srivenkatstock@gmail.com')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD','Sri@196768')

staff_mails = ['srivenkatreddy28@gmail.com','srivenkatreddy2208@gmail.com','srivenkatreddy28@gmail.com','srivenkatreddy2208@gmail.com']

warnings = {
    1: "Warning! You can take only one more day leave for ML class.",
    2: "Warning! You can take only one more day leave for DSA class.",
    3: "Warning! You can take only one more day leave for DBMS class.",
    4: "Warning! You can take only one more day leave for PYTHON class."
}

def save_file():
    data.save(r'/Users/srivenkatreddy/Documents/attendence.xlsx')
    print("Attendance data saved!!")

def send_mail(to_mail,subject,message):
    try:
        server = smtplib.SMTP('smtp@gmail.com',587)
        server.starttls()
        server.login(FROM_EMAIL,EMAIL_PASSWORD)

        email = MIMEMultipart()
        email['From'] = FROM_EMAIL
        email['T0'] = to_mail
        email['Subject'] = subject
        email.attach(MIMEText(message,'plain'))

        server.sendmail(FROM_EMAIL,to_mail,email.as_string())
        server.quit()
        print(f'Email sent to {to_mail} successfully')
    except Exception as e:
        print(f'Error sending Email to {to_mail} : {e}')

def check(no_of_days,row_num,subject_code):
    global std_mail_id,lack_attd,l3

    subject = {1:"ML",2:"DSA",3:"DBMS",4:"PYTHON"}.get(subject_code)

    for i,days in enumerate(no_of_days):
        if days==2:
            email = sheet.cell(row=row_num[i],column=2).value
            std_mail_id.append(email)
            send_mail(email,"Attendence warning",warnings[subject_code])
        elif days>2:
            row_num = sheet.cell(row=row_num[i],column=1).value
            lack_attd += f"roll_num"
            l3.append(sheet.cell(row=row_num[i],column=2).value)
    if lack_attd:
        staff_message = f"The following students have lack of attendance in {subject}: {l2.strip()}"
        send_mail(staff_mails[subject_code - 1], "Attendance Alert", staff_message)
        l2 = ""  # Reset after processing
        l3 = []
